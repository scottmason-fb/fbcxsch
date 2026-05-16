#!/usr/bin/env python3
"""
CX Schedule Automation Script
==============================
Run this script every Monday morning to:
  1. Read the current week's schedule from the Excel/Google Sheets workbook
  2. Print a formatted summary to the terminal
  3. (Optional) Post the summary to a Slack channel via webhook

Usage:
    python cx_schedule_automation.py                          # uses default file
    python cx_schedule_automation.py --file path/to/file.xlsx
    python cx_schedule_automation.py --week 2026-05-18        # specific Monday
    python cx_schedule_automation.py --slack-webhook URL      # post to Slack

Setup:
    pip install openpyxl requests
"""

import argparse
import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ─── CONFIG ──────────────────────────────────────────────────────────────────

DEFAULT_FILE    = "cx_schedule.xlsx"
ROSTER_SHEET    = "Roster"
TIMEOFF_SHEET   = "Time Off"
SCHEDULE_SHEET  = "Weekly Schedule"

# Column indices (1-based) in the Roster sheet
ROSTER_AGENT_COL   = 1   # A
ROSTER_TEAM_COL    = 2   # B
ROSTER_HOURS_COL   = 11  # K
ROSTER_DAY_START   = 4   # D = Monday ... J = Sunday (cols 4–10)

# Column indices in the Time Off sheet
TO_AGENT_COL    = 3   # C
TO_TEAM_COL     = 4   # D
TO_START_COL    = 5   # E
TO_END_COL      = 6   # F
TO_TYPE_COL     = 8   # H
TO_STATUS_COL   = 9   # I

DATA_START_ROW  = 4   # first data row in both sheets

WEEKDAY_MAP = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu",
               5: "Fri", 6: "Sat", 7: "Sun"}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def last_monday(ref: datetime.date) -> datetime.date:
    """Return the most recent Monday on or before ref."""
    return ref - datetime.timedelta(days=ref.weekday())


def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""


def to_date(v) -> datetime.date | None:
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def week_dates(monday: datetime.date) -> list[datetime.date]:
    return [monday + datetime.timedelta(days=i) for i in range(7)]


# ─── READ ROSTER ─────────────────────────────────────────────────────────────

def read_roster(wb) -> dict:
    """Returns {agent_name: {team, hours, schedule: {weekday_int: shift_str}}}"""
    ws = wb[ROSTER_SHEET]
    roster = {}
    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = cell_val(ws, row, ROSTER_AGENT_COL)
        if not name:
            continue
        team  = cell_val(ws, row, ROSTER_TEAM_COL)
        hours = cell_val(ws, row, ROSTER_HOURS_COL) or 0
        schedule = {}
        for offset in range(7):   # 0=Mon ... 6=Sun
            shift = cell_val(ws, row, ROSTER_DAY_START + offset)
            schedule[offset + 1] = str(shift) if shift else "OFF"
        roster[str(name)] = {"team": team, "hours": hours, "schedule": schedule}
    return roster


# ─── READ TIME OFF ───────────────────────────────────────────────────────────

def read_approved_timeoff(wb) -> list[dict]:
    """Returns list of approved time-off entries."""
    ws = wb[TIMEOFF_SHEET]
    approved = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        agent  = cell_val(ws, row, TO_AGENT_COL)
        status = str(cell_val(ws, row, TO_STATUS_COL)).strip()
        if not agent or status.lower() != "approved":
            continue
        start = to_date(ws.cell(row=row, column=TO_START_COL).value)
        end   = to_date(ws.cell(row=row, column=TO_END_COL).value)
        rtype = cell_val(ws, row, TO_TYPE_COL) or "Time Off"
        if start and end:
            approved.append({"agent": str(agent), "start": start,
                             "end": end, "type": str(rtype)})
    return approved


# ─── BUILD SCHEDULE ──────────────────────────────────────────────────────────

def build_schedule(roster: dict, timeoff: list[dict],
                   monday: datetime.date) -> dict:
    """
    Returns {agent_name: {date: display_string}} for the given week.
    """
    dates  = week_dates(monday)
    result = {}

    for agent, info in roster.items():
        week = {}
        for d in dates:
            weekday = d.isoweekday()  # 1=Mon … 7=Sun

            # Check for approved time off
            off_entry = next(
                (t for t in timeoff
                 if t["agent"] == agent and t["start"] <= d <= t["end"]),
                None
            )
            if off_entry:
                week[d] = f"[{off_entry['type']}]"
            else:
                week[d] = info["schedule"].get(weekday, "OFF")
        result[agent] = week
    return result


# ─── FORMAT SUMMARY ──────────────────────────────────────────────────────────

def format_summary(schedule: dict, roster: dict,
                   monday: datetime.date) -> str:
    dates    = week_dates(monday)
    day_hdrs = [f"{d.strftime('%a')} {d.strftime('%-m/%-d')}" for d in dates]

    lines = []
    lines.append("=" * 72)
    lines.append(f"  CX TEAM WEEKLY SCHEDULE — Week of {monday.strftime('%B %-d, %Y')}")
    lines.append("=" * 72)

    for team in ["Support", "Retail"]:
        lines.append(f"\n── {team.upper()} TEAM " + "─" * (60 - len(team)))
        team_agents = {a: info for a, info in roster.items()
                       if info["team"] == team}

        # Header row
        col_w = 14
        hdr_row = f"  {'Agent':<20}" + "".join(f"{h:^{col_w}}" for h in day_hdrs)
        lines.append(hdr_row)
        lines.append("  " + "-" * (20 + col_w * 7))

        # Agent rows
        available_by_day: dict[datetime.date, int] = {d: 0 for d in dates}
        for agent in sorted(team_agents.keys()):
            row = f"  {agent:<20}"
            for d in dates:
                val = schedule.get(agent, {}).get(d, "—")
                is_off = val.startswith("[") or val == "OFF"
                if not is_off:
                    available_by_day[d] += 1
                row += f"{val:^{col_w}}"
            lines.append(row)

        # Coverage row
        lines.append("  " + "-" * (20 + col_w * 7))
        cov_row = f"  {'Available':<20}"
        for d in dates:
            cov_row += f"{available_by_day[d]:^{col_w}}"
        lines.append(cov_row)

    # Pending time-off reminder
    lines.append("\n" + "─" * 72)
    lines.append("  ⚠  PENDING TIME-OFF REQUESTS — review and approve/deny:")
    lines.append("─" * 72)

    return "\n".join(lines)


def format_pending_reminder(wb, monday: datetime.date) -> list[str]:
    ws = wb[TIMEOFF_SHEET]
    end_of_week = monday + datetime.timedelta(days=6)
    pending = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        agent  = cell_val(ws, row, TO_AGENT_COL)
        status = str(cell_val(ws, row, TO_STATUS_COL)).strip()
        if not agent or status.lower() != "pending":
            continue
        start = to_date(ws.cell(row=row, column=TO_START_COL).value)
        end   = to_date(ws.cell(row=row, column=TO_END_COL).value)
        rtype = cell_val(ws, row, TO_TYPE_COL) or "Time Off"
        if start:
            pending.append(
                f"  • {agent:<22} {rtype:<12} "
                f"{start.strftime('%-m/%-d')}–{end.strftime('%-m/%-d') if end else '?'}"
            )
    return pending if pending else ["  (none)"]


# ─── SLACK ───────────────────────────────────────────────────────────────────

def post_to_slack(webhook_url: str, text: str) -> bool:
    try:
        import urllib.request
        payload = json.dumps({"text": f"```{text}```"}).encode()
        req = urllib.request.Request(
            webhook_url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status == 200
    except Exception as e:
        print(f"  ✗ Slack post failed: {e}")
        return False


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="CX Schedule Automation — generate weekly schedule summary"
    )
    parser.add_argument("--file", default=DEFAULT_FILE,
                        help="Path to the scheduling .xlsx file")
    parser.add_argument("--week", default=None,
                        help="Monday date for the target week (YYYY-MM-DD). "
                             "Defaults to the current or upcoming Monday.")
    parser.add_argument("--slack-webhook", default=None,
                        help="Slack incoming webhook URL to post the summary")
    args = parser.parse_args()

    # ── Resolve file ──
    fpath = Path(args.file)
    if not fpath.exists():
        print(f"Error: file not found: {fpath}")
        sys.exit(1)

    # ── Resolve week ──
    if args.week:
        try:
            monday = datetime.date.fromisoformat(args.week)
        except ValueError:
            print(f"Error: --week must be YYYY-MM-DD, got '{args.week}'")
            sys.exit(1)
        if monday.weekday() != 0:
            print(f"Warning: {monday} is not a Monday. Adjusting to nearest Monday.")
            monday = last_monday(monday)
    else:
        today  = datetime.date.today()
        monday = last_monday(today)
        # If today is Sunday, bump to next Monday
        if today.weekday() == 6:
            monday += datetime.timedelta(days=1)

    print(f"Loading workbook: {fpath}")
    print(f"Building schedule for week of: {monday}\n")

    # ── Load data ──
    wb      = openpyxl.load_workbook(str(fpath), data_only=True)
    roster  = read_roster(wb)
    timeoff = read_approved_timeoff(wb)

    if not roster:
        print("⚠  No agents found in Roster sheet. Please add agents first.")
        sys.exit(1)

    print(f"  Loaded {len(roster)} agents, {len(timeoff)} approved time-off requests\n")

    # ── Build & print ──
    schedule = build_schedule(roster, timeoff, monday)
    summary  = format_summary(schedule, roster, monday)
    pending  = format_pending_reminder(wb, monday)

    full_output = summary + "\n" + "\n".join(pending) + "\n" + "=" * 72 + "\n"
    print(full_output)

    # ── Slack ──
    if args.slack_webhook:
        print("Posting to Slack...", end=" ")
        ok = post_to_slack(args.slack_webhook, full_output)
        print("✓ Done" if ok else "✗ Failed")


if __name__ == "__main__":
    main()
